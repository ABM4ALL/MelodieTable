from typing import Generator, List, Tuple, Union
import openpyxl
import csv

TableValues = Tuple[List[str],
                    Generator[List[Union[str, int, float]], None, None]]


class TableReader:
    def __init__(self, file_name: str, header=0, text_encoding="utf-8") -> None:
        self.file_name = file_name
        self.type = ""
        self.text_encoding = text_encoding
        self.header = header
        if file_name.endswith(".csv"):
            self.type = "csv"
        elif file_name.endswith((".xls", ".xlsx")):
            self.type = "excel"
        else:
            raise NotImplementedError
        self.read_methods = {
            "csv": self._read_csv,
            "excel": self._read_excel
        }

    def read(self):
        return self.read_methods[self.type]()

    def _read_csv(self) -> TableValues:
        f = open(self.file_name, encoding=self.text_encoding)
        reader = csv.reader(f)
        current_row = 0
        header: List[str] = []
        while current_row <= self.header:
            header = next(reader)
            current_row += 1

        def row_iter():
            for row_data in reader:
                yield row_data
            f.close()
        return header, row_iter()

    def _read_excel(self) -> TableValues:
        def excel_max_row(sheet):
            i = sheet.max_row
            real_max_row = 0
            while i > 0:
                row_dict = {i.value for i in sheet[i]}
                if row_dict == {None}:
                    i = i-1
                else:
                    real_max_row = i
                    break

            return real_max_row

        workbook = openpyxl.load_workbook(self.file_name)
        table = workbook.active
        rows = excel_max_row(table)
        cols = table.max_column
        header = [table.cell(self.header + 1, col +
                             1).value for col in range(cols)]

        def row_iter():
            for row_index in range(self.header+1, rows):
                row = [table.cell(
                    row_index + 1, col + 1).value for col in range(cols)]
                yield row
        return header, row_iter()


class TableWriter:
    def __init__(self, file_name: str, header=0, text_encoding="utf-8") -> None:
        self.file_name = file_name
        self.type = ""
        self.text_encoding = text_encoding
        self.header = header
        if file_name.endswith(".csv"):
            self.type = "csv"
        elif file_name.endswith((".xls", ".xlsx")):
            self.type = "excel"
        else:
            raise NotImplementedError
        self.write_methods = {
            "csv": self._write_csv,
            "excel": self._write_excel
        }

    def write(self):
        return self.write_methods[self.type]()

    def _write_csv(self):
        file = open(self.file_name, "w",
                    encoding=self.text_encoding, newline='')
        writer = csv.writer(file)
        current_row = 1

        def row_writer():
            nonlocal current_row
            while 1:
                try:
                    data = yield
                    # for item in data:
                    #     file.write(str(item))
                    #     file.write(",")
                    # file.write("\n")
                    writer.writerow(data)
                    current_row += 1
                except GeneratorExit:
                    file.close()
                    return
        w = row_writer()
        next(w)
        return w

    def _write_excel(self):
        wb = openpyxl.Workbook()
        # 获取当前活跃的worksheet,默认就是第一个worksheet
        ws = wb.active
        current_row = 1

        def row_writer():
            nonlocal current_row
            while 1:
                try:
                    data = yield
                    for j, val in enumerate(data):
                        ws.cell(row=current_row, column=j + 1).value = val
                    current_row += 1
                except GeneratorExit:
                    wb.save(filename=self.file_name)
                    return
        # 保存表格
        w = row_writer()
        next(w)
        return w

if __name__ == "__main__":
    r = TableReader(
        r"F:\Developing\melodie-table\tests\data\pyam_tutorial_data.xlsx")
    header, rows = r._read_excel()
    for r in rows:
        print(r)
